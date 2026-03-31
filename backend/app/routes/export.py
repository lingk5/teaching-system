"""
数据导出模块 - 教学效果监督系统
支持：学生名单、成绩报表、考勤统计、预警报告导出
格式：Excel (.xlsx) 带样式美化
"""
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required
from sqlalchemy import func, and_, or_, case
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import traceback

from ..models import db, Student, Course, Class
from ..models.data import Attendance, Homework, Quiz, FinalScore, Interaction
from ..models.warning import Warning
from ..services.weight_config import WeightConfig
from ..utils.permissions import current_user_can, accessible_course_ids, can_access_course

export_bp = Blueprint('export', __name__)

# 样式配置
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def auto_adjust_column_width(worksheet):
    """自动调整列宽"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def apply_styles(worksheet):
    """应用表头样式和边框"""
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.row == 1:  # 表头行
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')

@export_bp.route('/students', methods=['GET'])
@jwt_required()
def export_students():
    """
    导出学生名单
    """
    try:
        if not current_user_can('export_reports'):
            return jsonify({'success': False, 'message': '助教无权限导出报表'}), 403

        allowed_course_ids = accessible_course_ids()

        course_id = request.args.get('course_id', type=int)
        class_id = request.args.get('class_id', type=int)
        format_type = request.args.get('format', 'xlsx')

        if course_id and not can_access_course(course_id):
            return jsonify({'success': False, 'message': '无权导出该课程数据'}), 403

        # 构建查询
        query = db.session.query(
            Student,
            Class.name.label('class_name')
        ).join(
            Class, Student.class_id == Class.id
        )

        # 如果指定了课程，只导出该课程的学生
        if course_id:
            query = query.join(
                Course, Class.course_id == Course.id
            ).filter(
                Course.id == course_id
            )
            course = Course.query.filter_by(id=course_id).first()
            if not course:
                return jsonify({'success': False, 'message': '课程不存在'}), 404
            filename_suffix = f"_{course.name}"
        else:
            query = query.join(
                Course, Class.course_id == Course.id
            )
            query = query.filter(Course.id.in_(allowed_course_ids))
            filename_suffix = "_全部"

        if class_id:
            query = query.filter(Student.class_id == class_id)

        results = query.distinct().all()

        if not results:
            return jsonify({'success': False, 'message': '无数据可导出'}), 404

        # 构建DataFrame
        data = []
        for student, class_name in results:
            data.append({
                '学号': student.student_no,
                '姓名': student.name,
                '性别': student.gender or '',
                '班级': class_name,
                '创建日期': student.created_at.strftime('%Y-%m-%d') if student.created_at else ''
            })

        df = pd.DataFrame(data)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"学生名单{filename_suffix}_{timestamp}.xlsx"

        if format_type == 'csv':
            output = BytesIO()
            df.to_csv(output, index=False, encoding='utf-8-sig')
            output.seek(0)
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename.replace('.xlsx', '.csv')
            )
        else:
            # Excel格式，带样式
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='学生名单', index=False)
                worksheet = writer.sheets['学生名单']
                apply_styles(worksheet)
                auto_adjust_column_width(worksheet)

            output.seek(0)
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

    except Exception as e:
        current_app.logger.error(f"导出学生失败: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500

@export_bp.route('/scores', methods=['GET'])
@jwt_required()
def export_scores():
    """
    导出成绩报表
    """
    try:
        if not current_user_can('export_reports'):
            return jsonify({'success': False, 'message': '助教无权限导出报表'}), 403

        allowed_course_ids = accessible_course_ids()

        course_id = request.args.get('course_id', type=int)
        class_id = request.args.get('class_id', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        if course_id and not can_access_course(course_id):
            return jsonify({'success': False, 'message': '无权导出该课程数据'}), 403

        attendance_score_case = case(
            (Attendance.status == 'present', 100),
            (Attendance.status == 'late', 80),
            (Attendance.status == 'leave', 60),
            (Attendance.status == 'absent', 0),
            else_=0
        )

        attendance_query = db.session.query(
            Attendance.student_id.label('student_id'),
            Attendance.course_id.label('course_id'),
            func.avg(attendance_score_case).label('attendance_avg')
        )
        if start_date:
            attendance_query = attendance_query.filter(Attendance.date >= start_date)
        if end_date:
            attendance_query = attendance_query.filter(Attendance.date <= end_date)
        attendance_subquery = attendance_query.group_by(
            Attendance.student_id, Attendance.course_id
        ).subquery()

        homework_query = db.session.query(
            Homework.student_id.label('student_id'),
            Homework.course_id.label('course_id'),
            func.avg(Homework.score / func.nullif(Homework.max_score, 0) * 100).label('homework_avg')
        )
        if start_date:
            homework_query = homework_query.filter(Homework.created_at >= start_date)
        if end_date:
            homework_query = homework_query.filter(Homework.created_at <= end_date)
        homework_subquery = homework_query.group_by(
            Homework.student_id, Homework.course_id
        ).subquery()

        quiz_query = db.session.query(
            Quiz.student_id.label('student_id'),
            Quiz.course_id.label('course_id'),
            func.avg(Quiz.score / func.nullif(Quiz.max_score, 0) * 100).label('quiz_avg')
        )
        if start_date:
            quiz_query = quiz_query.filter(Quiz.created_at >= start_date)
        if end_date:
            quiz_query = quiz_query.filter(Quiz.created_at <= end_date)
        quiz_subquery = quiz_query.group_by(
            Quiz.student_id, Quiz.course_id
        ).subquery()

        final_exam_query = db.session.query(
            FinalScore.student_id.label('student_id'),
            FinalScore.course_id.label('course_id'),
            func.avg(FinalScore.score / func.nullif(FinalScore.max_score, 0) * 100).label('final_exam_avg')
        )
        if start_date:
            final_exam_query = final_exam_query.filter(FinalScore.created_at >= start_date)
        if end_date:
            final_exam_query = final_exam_query.filter(FinalScore.created_at <= end_date)
        final_exam_subquery = final_exam_query.group_by(
            FinalScore.student_id, FinalScore.course_id
        ).subquery()

        interaction_query = db.session.query(
            Interaction.student_id.label('student_id'),
            Interaction.course_id.label('course_id'),
            func.sum(Interaction.count).label('interaction_total')
        )
        if start_date:
            interaction_query = interaction_query.filter(Interaction.date >= start_date)
        if end_date:
            interaction_query = interaction_query.filter(Interaction.date <= end_date)
        interaction_subquery = interaction_query.group_by(
            Interaction.student_id, Interaction.course_id
        ).subquery()

        query = db.session.query(
            Student.id.label('student_id'),
            Student.student_no,
            Student.name,
            Course.id.label('course_id'),
            Course.name.label('course_name'),
            Class.name.label('class_name'),
            attendance_subquery.c.attendance_avg,
            homework_subquery.c.homework_avg,
            quiz_subquery.c.quiz_avg,
            final_exam_subquery.c.final_exam_avg,
            interaction_subquery.c.interaction_total,
        ).join(
            Class, Student.class_id == Class.id
        ).join(
            Course, Class.course_id == Course.id
        ).outerjoin(
            attendance_subquery,
            and_(
                attendance_subquery.c.student_id == Student.id,
                attendance_subquery.c.course_id == Course.id,
            )
        ).outerjoin(
            homework_subquery,
            and_(
                homework_subquery.c.student_id == Student.id,
                homework_subquery.c.course_id == Course.id,
            )
        ).outerjoin(
            quiz_subquery,
            and_(
                quiz_subquery.c.student_id == Student.id,
                quiz_subquery.c.course_id == Course.id,
            )
        ).outerjoin(
            final_exam_subquery,
            and_(
                final_exam_subquery.c.student_id == Student.id,
                final_exam_subquery.c.course_id == Course.id,
            )
        ).outerjoin(
            interaction_subquery,
            and_(
                interaction_subquery.c.student_id == Student.id,
                interaction_subquery.c.course_id == Course.id,
            )
        ).filter(
            Course.id.in_(allowed_course_ids)
        )

        if course_id:
            query = query.filter(Course.id == course_id)
        if class_id:
            query = query.filter(Student.class_id == class_id)

        results = query.order_by(Course.name.asc(), Class.name.asc(), Student.student_no.asc()).all()

        if not results:
            return jsonify({'success': False, 'message': '无成绩数据可导出'}), 404

        # 构建数据（使用统一的权重配置）
        col_titles = WeightConfig.get_score_column_titles()
        data = []
        for row in results:
            attendance = float(row.attendance_avg) if row.attendance_avg is not None else None
            homework = float(row.homework_avg) if row.homework_avg is not None else None
            quiz = float(row.quiz_avg) if row.quiz_avg is not None else None
            final_exam = float(row.final_exam_avg) if row.final_exam_avg is not None else None
            interaction = min(float(row.interaction_total) * 10, 100) if row.interaction_total is not None else None

            metrics = {
                'attendance': attendance,
                'homework': homework,
                'quiz': quiz,
                'final_exam': final_exam,
                'interaction': interaction,
            }
            score_details = WeightConfig.calculate_score_details(metrics)
            composite = score_details['comprehensive_score']

            # 评定等级
            if composite >= 90:
                grade = '优秀'
            elif composite >= 80:
                grade = '良好'
            elif composite >= 70:
                grade = '中等'
            elif composite >= 60:
                grade = '及格'
            else:
                grade = '不及格'

            data.append({
                '学号':                    row.student_no,
                '姓名':                    row.name,
                '班级':                    row.class_name,
                '课程':                    row.course_name,
                col_titles['attendance']:  round(attendance, 1) if attendance is not None else '',
                col_titles['homework']:    round(homework, 1) if homework is not None else '',
                col_titles['quiz']:        round(quiz, 1) if quiz is not None else '',
                col_titles['final_exam']:  round(final_exam, 1) if final_exam is not None else '',
                col_titles['interaction']: round(interaction, 1) if interaction is not None else '',
                '综合分':                  round(composite, 1),
                '等级':                    grade,
            })

        df = pd.DataFrame(data)
        df = df.sort_values('综合分', ascending=False)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        course = db.session.get(Course, course_id) if course_id else None
        course_name = course.name if course else "全部课程"
        filename = f"成绩报表_{course_name}_{timestamp}.xlsx"

        # 创建Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='成绩报表', index=False)
            worksheet = writer.sheets['成绩报表']

            apply_styles(worksheet)
            auto_adjust_column_width(worksheet)

            # 为等级列添加条件格式（颜色）
            grade_col = None
            for idx, col_name in enumerate(df.columns, 1):
                if col_name == '等级':
                    grade_col = idx
                    break

            if grade_col:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=grade_col)
                    val = cell.value
                    if val == '优秀':
                        cell.font = Font(color="006100")
                        cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
                    elif val == '不及格':
                        cell.font = Font(color="9C0006")
                        cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"导出成绩失败: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500

@export_bp.route('/attendance', methods=['GET'])
@jwt_required()
def export_attendance():
    """
    导出考勤统计
    """
    try:
        if not current_user_can('export_reports'):
            return jsonify({'success': False, 'message': '助教无权限导出报表'}), 403

        allowed_course_ids = accessible_course_ids()

        course_id = request.args.get('course_id', type=int)

        # 默认本月
        today = datetime.now().date()
        start_date = request.args.get('start_date', today.replace(day=1).strftime('%Y-%m-%d'))
        end_date = request.args.get('end_date', today.strftime('%Y-%m-%d'))

        # 查询考勤记录
        query = db.session.query(
            Student.student_no,
            Student.name,
            Class.name.label('class_name'),
            Course.name.label('course_name'),
            Attendance.date,
            Attendance.status,
            Attendance.remark
        ).join(
            Attendance, Student.id == Attendance.student_id
        ).join(
            Course, Attendance.course_id == Course.id
        ).join(
            Class, Student.class_id == Class.id
        ).filter(
            Attendance.date.between(start_date, end_date)
        )
        
        query = query.filter(Course.id.in_(allowed_course_ids))

        if course_id:
            query = query.filter(Attendance.course_id == course_id)

        results = query.order_by(Attendance.date.desc()).all()

        if not results:
            return jsonify({'success': False, 'message': '该时间段无考勤记录'}), 404

        # 构建数据
        data = []
        status_map = {'present': '出勤', 'late': '迟到', 'absent': '缺勤', 'leave': '请假'}

        for row in results:
            data.append({
                '日期': row.date.strftime('%Y-%m-%d'),
                '学号': row.student_no,
                '姓名': row.name,
                '班级': row.class_name,
                '课程': row.course_name,
                '考勤状态': status_map.get(row.status, row.status),
                '备注': row.remark or ''
            })

        df = pd.DataFrame(data)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"考勤统计_{start_date}_to_{end_date}_{timestamp}.xlsx"

        # 创建Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: 明细
            df.to_excel(writer, sheet_name='考勤明细', index=False)
            worksheet1 = writer.sheets['考勤明细']
            apply_styles(worksheet1)
            auto_adjust_column_width(worksheet1)

            # Sheet 2: 汇总统计
            summary_data = []
            if '课程' in df.columns:
                for course in df['课程'].unique():
                    course_df = df[df['课程'] == course]
                    total = len(course_df)
                    present = len(course_df[course_df['考勤状态'] == '出勤'])
                    late = len(course_df[course_df['考勤状态'] == '迟到'])
                    absent = len(course_df[course_df['考勤状态'] == '缺勤'])
                    rate = round(present / total * 100, 1) if total > 0 else 0

                    summary_data.append({
                        '课程': course,
                        '总人次': total,
                        '出勤': present,
                        '迟到': late,
                        '缺勤': absent,
                        '出勤率': f'{rate}%'
                    })

            if summary_data:
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
                worksheet2 = writer.sheets['汇总统计']
                apply_styles(worksheet2)
                auto_adjust_column_width(worksheet2)

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"导出考勤失败: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500

@export_bp.route('/warnings', methods=['GET'])
@jwt_required()
def export_warnings():
    """
    导出预警报告
    """
    try:
        if not current_user_can('export_reports'):
            return jsonify({'success': False, 'message': '助教无权限导出报表'}), 403

        allowed_course_ids = accessible_course_ids()

        level = request.args.get('level')
        status = request.args.get('status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        course_id = request.args.get('course_id', type=int)

        if course_id and not can_access_course(course_id):
            return jsonify({'success': False, 'message': '无权导出该课程数据'}), 403

        # 构建查询
        query = db.session.query(
            Warning,
            Student.student_no,
            Student.name,
            Course.name.label('course_name')
        ).join(
            Student, Warning.student_id == Student.id
        ).join(
            Course, Warning.course_id == Course.id
        )
        
        query = query.filter(Course.id.in_(allowed_course_ids))
        if course_id:
            query = query.filter(Course.id == course_id)

        if level:
            query = query.filter(Warning.level == level)
        if status:
            query = query.filter(Warning.status == status)
        if start_date:
            query = query.filter(Warning.created_at >= start_date)
        if end_date:
            query = query.filter(Warning.created_at <= end_date)

        results = query.order_by(Warning.created_at.desc()).all()

        if not results:
            return jsonify({'success': False, 'message': '无预警记录可导出'}), 404

        # 构建数据
        data = []
        level_map = {'red': '红色', 'orange': '橙色', 'yellow': '黄色'}
        status_map = {'active': '未处理', 'processed': '已处理', 'ignored': '已忽略', 'following': '跟进中'}

        for warning, student_no, name, course_name in results:
            data.append({
                '预警日期': warning.created_at.strftime('%Y-%m-%d %H:%M'),
                '学号': student_no,
                '姓名': name,
                '课程': course_name,
                '预警等级': level_map.get(warning.level, warning.level),
                '预警类型': warning.type,
                '预警原因': warning.reason,
                '处理建议': warning.suggestion or '',
                '当前状态': status_map.get(warning.status, warning.status),
                '处理时间': warning.handled_at.strftime('%Y-%m-%d %H:%M') if warning.handled_at else '',
                '处理备注': warning.handle_note or ''
            })

        df = pd.DataFrame(data)

        # 生成文件名
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        level_suffix = f"_{level_map.get(level, '全部')}"
        filename = f"预警报告{level_suffix}_{timestamp}.xlsx"

        # 创建Excel
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='预警明细', index=False)
            worksheet = writer.sheets['预警明细']

            apply_styles(worksheet)
            auto_adjust_column_width(worksheet)

            # 为预警等级添加颜色
            level_col = None
            for idx, col_name in enumerate(df.columns, 1):
                if col_name == '预警等级':
                    level_col = idx
                    break

            if level_col:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=level_col)
                    val = cell.value
                    if val == '红色':
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.fill = PatternFill(start_color="DC3545", fill_type="solid")
                    elif val == '橙色':
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.fill = PatternFill(start_color="FD7E14", fill_type="solid")
                    elif val == '黄色':
                        cell.font = Font(color="000000", bold=True)
                        cell.fill = PatternFill(start_color="FFC107", fill_type="solid")

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.error(f"导出预警失败: {str(e)}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'}), 500

@export_bp.route('/template/<type>', methods=['GET'])
def download_template(type):
    """
    下载导入模板
    type: students / scores / attendance
    """
    try:
        output = BytesIO()
        filename = f'{type}_template.xlsx'

        if type == 'students':
            df = pd.DataFrame({
                'student_no': ['2024001', '2024002'],
                'name': ['张三', '李四'],
                'gender': ['男', '女'],
                'class_name': ['计算机9班', '计算机9班']
            })
            instructions = [
                'student_no: 学号 (必填)',
                'name: 姓名 (必填)',
                'gender: 性别 (可选)',
                'class_name: 班级名称 (必填，如班级不存在将自动创建)'
            ]

        elif type == 'scores':
            df = pd.DataFrame({
                'student_no': ['2024001', '2024002', '2024003'],
                'title': ['期中作业', '第一次测验', 'Python程序设计期末'],
                'score': [85, 90, 88],
                'max_score': [100, 100, 100],
                'type': ['homework', 'quiz', 'final_exam'],
                'duration': ['', 60, 120]
            })
            instructions = [
                'student_no: 学号 (必填)',
                'title: 标题 (必填)',
                'score: 分数 (必填)',
                'max_score: 满分 (默认为100)',
                'type: 类型 (homework/quiz/final_exam，默认为homework)',
                'duration: 时长(分钟)，type 为 quiz/final_exam 时有效'
            ]

        elif type == 'attendance':
            df = pd.DataFrame({
                'student_no': ['2024001', '2024002'],
                'date': ['2026-03-01', '2026-03-01'],
                'status': ['present', 'absent'],
                'remark': ['', '病假']
            })
            instructions = [
                'student_no: 学号 (必填)',
                'date: 日期 (必填, YYYY-MM-DD)',
                'status: 状态 (必填, present/absent/late/leave)',
                'remark: 备注 (可选)'
            ]
        else:
            return jsonify({'success': False, 'message': '未知模板类型'}), 400

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='数据模板', index=False)
            worksheet = writer.sheets['数据模板']
            apply_styles(worksheet)
            auto_adjust_column_width(worksheet)

            # 添加说明页
            inst_df = pd.DataFrame({'字段说明': instructions})
            inst_df.to_excel(writer, sheet_name='填写说明', index=False)
            writer.sheets['填写说明'].column_dimensions['A'].width = 60

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'success': False, 'message': f'生成模板失败: {str(e)}'}), 500
        if course_id and not can_access_course(course_id):
            return jsonify({'success': False, 'message': '无权导出该课程数据'}), 403
